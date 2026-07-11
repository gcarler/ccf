"""
Reportes de Evangelismo — PDF de asistencia, Excel de asistencia,
y resumen por estrategia.

Endpoints:
  GET /api/evangelism/reports/group/{grupo_id}/attendance-pdf
  GET /api/evangelism/reports/group/{grupo_id}/attendance-excel
  GET /api/evangelism/reports/strategy/{strategy_id}/summary
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from backend import models
from backend.api.evangelism_shared import (
    ATTENDED_STATES,
    is_absent_status,
    is_attended_status,
    is_excused_status,
    session_read_only_options,
    session_read_value,
)
from backend.core.database import get_db
from backend.core.permissions import require_active_user
from backend.core.tenant import require_user_sede_id

logger = logging.getLogger(__name__)

router = APIRouter()

# ──────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────

def _get_group_or_404(db: Session, grupo_id: UUID) -> models.GrupoEvangelismo:
    grupo = db.query(models.GrupoEvangelismo).filter(
        models.GrupoEvangelismo.id == grupo_id,
        models.GrupoEvangelismo.deleted_at.is_(None),
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    return grupo


def _get_leader_name(db: Session, grupo: models.GrupoEvangelismo) -> str:
    if grupo.lider_persona_id is None:
        return "Sin líder asignado"
    persona = db.query(models.Persona).filter(
        models.Persona.id == grupo.lider_persona_id
    ).first()
    if persona:
        return f"{persona.first_name} {persona.last_name}".strip()
    return "Líder (no encontrado)"


def _count_participants(db: Session, grupo_id: UUID) -> int:
    return db.query(func.count(models.ParticipanteGrupo.id)).filter(
        models.ParticipanteGrupo.grupo_id == grupo_id,
        models.ParticipanteGrupo.activo,
    ).scalar() or 0


def _build_session_rows(
    db: Session, grupo_id: UUID
) -> list[dict]:
    """Return one dict per session with attendance stats."""
    sessions = (
        db.query(models.SesionGrupo)
        .options(session_read_only_options(db))
        .filter(
            models.SesionGrupo.grupo_id == grupo_id,
            models.SesionGrupo.deleted_at.is_(None),
        )
        .order_by(models.SesionGrupo.fecha_sesion.asc())
        .all()
    )

    total_participants = _count_participants(db, grupo_id)
    session_ids = [sesion.id for sesion in sessions]
    asistencia_by_session = defaultdict(list)
    if session_ids:
        for asistencia in (
            db.query(models.Asistencia)
            .filter(
                models.Asistencia.sesion_id.in_(session_ids),
                models.Asistencia.deleted_at.is_(None),
            )
            .all()
        ):
            asistencia_by_session[asistencia.sesion_id].append(asistencia)
    rows: list[dict] = []
    for sesion in sessions:
        asistencias = asistencia_by_session.get(sesion.id, [])

        asistentes = sum(1 for a in asistencias if is_attended_status(a.estado))
        ausentes = sum(1 for a in asistencias if is_absent_status(a.estado))
        excusas = sum(1 for a in asistencias if is_excused_status(a.estado))
        # porcentaje: asistentes / total_participants * 100
        if total_participants > 0:
            pct = round(asistentes / total_participants * 100, 1)
        else:
            pct = 0.0

        rows.append({
            "fecha": sesion.fecha_sesion.strftime("%d/%m/%Y") if sesion.fecha_sesion else "—",
            "tema": session_read_value(sesion, "tema_estudio")
            or session_read_value(sesion, "topic")
            or "Sin tema",
            "estado": session_read_value(sesion, "estado")
            or session_read_value(sesion, "status")
            or "—",
            "asistentes": asistentes,
            "ausentes": ausentes,
            "excusas": excusas,
            "total_participantes": total_participants,
            "pct_asistencia": pct,
        })
    return rows


# ──────────────────────────────────────────────────────────────────────
# PDF Report (reportlab)
# ──────────────────────────────────────────────────────────────────────

def _generate_attendance_pdf(
    grupo: models.GrupoEvangelismo,
    leader_name: str,
    rows: list[dict],
) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]

    story = []

    # Title
    story.append(Paragraph("Reporte de Asistencia — Grupo de Evangelismo", title_style))
    story.append(Spacer(1, 0.5 * cm))

    # Group info
    story.append(Paragraph(f"<b>Grupo:</b> {grupo.nombre}", normal_style))
    story.append(Paragraph(f"<b>Líder:</b> {leader_name}", normal_style))
    now_str = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    story.append(Paragraph(f"<b>Fecha de emisión:</b> {now_str}", normal_style))
    story.append(Spacer(1, 0.5 * cm))

    # Table
    if rows:
        table_data = [
            ["Fecha", "Tema", "Estado", "Asist.", "Aus.", "Exc.", "% Asist."],
        ]
        for r in rows:
            table_data.append([
                r["fecha"],
                r["tema"],
                r["estado"],
                str(r["asistentes"]),
                str(r["ausentes"]),
                str(r["excusas"]),
                f"{r['pct_asistencia']}%",
            ])

        col_widths = [2.2 * cm, 5.5 * cm, 2.0 * cm, 1.3 * cm, 1.3 * cm, 1.3 * cm, 2.0 * cm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ecf0f1")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("<i>No hay sesiones registradas para este grupo.</i>", normal_style))

    story.append(Spacer(1, 1.5 * cm))

    # Signature line
    story.append(Paragraph("_" * 50, normal_style))
    story.append(Paragraph(f"<b>Firma del Líder:</b> {leader_name}", normal_style))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("CCF — Sistema de Inteligencia Ministerial", normal_style))

    doc.build(story)
    buf.seek(0)
    return buf


@router.get("/reports/group/{grupo_id}/attendance-pdf")
def attendance_pdf(
    grupo_id: UUID,
    db: Session = Depends(get_db),
    current_user=Depends(require_active_user),
):
    """Genera PDF de asistencia del grupo con tabla de sesiones y línea de firma."""
    grupo = _get_group_or_404(db, grupo_id)
    user_sede = require_user_sede_id(db, current_user)
    if user_sede is not None and grupo.sede_id and str(grupo.sede_id) != str(user_sede):
        raise HTTPException(status_code=403, detail="Grupo no pertenece a tu sede")
    leader_name = _get_leader_name(db, grupo)
    rows = _build_session_rows(db, grupo_id)

    buf = _generate_attendance_pdf(grupo, leader_name, rows)

    filename = f"asistencia_grupo_{grupo_id}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──────────────────────────────────────────────────────────────────────
# Excel Report (openpyxl)
# ──────────────────────────────────────────────────────────────────────

def _generate_attendance_excel(
    grupo: models.GrupoEvangelismo,
    leader_name: str,
    rows: list[dict],
) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Header section
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Reporte de Asistencia — {grupo.nombre}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Líder:"
    ws["B2"] = leader_name
    ws["B2"].font = Font(bold=True)

    now_str = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    ws["A3"] = "Fecha de emisión:"
    ws["B3"] = now_str

    # Column headers
    headers = ["Fecha", "Tema", "Estado", "Asistentes", "Ausentes", "Excusas", "% Asistencia"]
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    start_row = 5
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, r in enumerate(rows):
        row_num = start_row + 1 + i
        values = [
            r["fecha"],
            r["tema"],
            r["estado"],
            r["asistentes"],
            r["ausentes"],
            r["excusas"],
            f"{r['pct_asistencia']}%",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [14, 30, 14, 12, 12, 12, 14]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = w

    # Signature row
    sig_row = start_row + 1 + len(rows) + 2
    ws.merge_cells(f"A{sig_row}:G{sig_row}")
    ws[f"A{sig_row}"] = f"Firma del Líder: {leader_name}"
    ws[f"A{sig_row}"].font = Font(italic=True)
    sig_row2 = sig_row + 1
    ws.merge_cells(f"A{sig_row2}:G{sig_row2}")
    ws[f"A{sig_row2}"] = "CCF — Sistema de Inteligencia Ministerial"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/reports/group/{grupo_id}/attendance-excel")
def attendance_excel(
    grupo_id: UUID,
    db: Session = Depends(get_db),
    current_user=Depends(require_active_user),
):
    """Genera Excel de asistencia del grupo con tabla de sesiones."""
    grupo = _get_group_or_404(db, grupo_id)
    user_sede = require_user_sede_id(db, current_user)
    if user_sede is not None and grupo.sede_id and str(grupo.sede_id) != str(user_sede):
        raise HTTPException(status_code=403, detail="Grupo no pertenece a tu sede")
    leader_name = _get_leader_name(db, grupo)
    rows = _build_session_rows(db, grupo_id)

    buf = _generate_attendance_excel(grupo, leader_name, rows)

    filename = f"asistencia_grupo_{grupo_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──────────────────────────────────────────────────────────────────────
# Strategy Summary
# ──────────────────────────────────────────────────────────────────────

@router.get("/reports/strategy/{strategy_id}/summary")
def strategy_summary(
    strategy_id: UUID,
    db: Session = Depends(get_db),
    current_user=Depends(require_active_user),
):
    """Resumen de todos los grupos de una estrategia de evangelismo."""
    user_sede = require_user_sede_id(db, current_user)
    estrategia_q = db.query(models.EstrategiaEvangelismo).filter(
        models.EstrategiaEvangelismo.id == strategy_id,
        models.EstrategiaEvangelismo.deleted_at.is_(None),
    )
    if user_sede is not None:
        estrategia_q = estrategia_q.filter(models.EstrategiaEvangelismo.sede_id == user_sede)
    estrategia = estrategia_q.first()
    if not estrategia:
        raise HTTPException(status_code=404, detail="Estrategia no encontrada")

    grupos_q = db.query(models.GrupoEvangelismo).filter(
        models.GrupoEvangelismo.estrategia_id == strategy_id,
        models.GrupoEvangelismo.deleted_at.is_(None),
    )
    if user_sede is not None:
        grupos_q = grupos_q.filter(models.GrupoEvangelismo.sede_id == user_sede)
    grupos = grupos_q.all()
    grupo_ids = [grupo.id for grupo in grupos]

    participantes_map = dict(
        db.query(models.ParticipanteGrupo.grupo_id, func.count(models.ParticipanteGrupo.id))
        .filter(models.ParticipanteGrupo.grupo_id.in_(grupo_ids) if grupo_ids else False)
        .filter(models.ParticipanteGrupo.activo, models.ParticipanteGrupo.deleted_at.is_(None))
        .group_by(models.ParticipanteGrupo.grupo_id)
        .all()
    )
    sesiones_map = dict(
        db.query(models.SesionGrupo.grupo_id, func.count(models.SesionGrupo.id))
        .filter(models.SesionGrupo.grupo_id.in_(grupo_ids) if grupo_ids else False)
        .filter(models.SesionGrupo.deleted_at.is_(None))
        .group_by(models.SesionGrupo.grupo_id)
        .all()
    )
    realizadas_map = dict(
        db.query(models.SesionGrupo.grupo_id, func.count(models.SesionGrupo.id))
        .filter(models.SesionGrupo.grupo_id.in_(grupo_ids) if grupo_ids else False)
        .filter(models.SesionGrupo.estado == "REALIZADA")
        .filter(models.SesionGrupo.deleted_at.is_(None))
        .group_by(models.SesionGrupo.grupo_id)
        .all()
    )
    present_by_group_session = defaultdict(dict)
    if grupo_ids:
        for grupo_id, sesion_id, total in (
            db.query(
                models.SesionGrupo.grupo_id,
                models.SesionGrupo.id,
                func.count(models.Asistencia.id),
            )
            .outerjoin(
                models.Asistencia,
                (models.Asistencia.sesion_id == models.SesionGrupo.id)
                & (models.Asistencia.estado.in_(ATTENDED_STATES))
                & (models.Asistencia.deleted_at.is_(None)),
            )
            .filter(
                models.SesionGrupo.grupo_id.in_(grupo_ids),
                models.SesionGrupo.deleted_at.is_(None),
            )
            .group_by(models.SesionGrupo.grupo_id, models.SesionGrupo.id)
            .all()
        ):
            present_by_group_session[grupo_id][sesion_id] = total

    grupos_resumen = []
    for grupo in grupos:
        leader_name = _get_leader_name(db, grupo)
        total_participantes = participantes_map.get(grupo.id, 0)
        total_sesiones = sesiones_map.get(grupo.id, 0)
        sesiones_realizadas = realizadas_map.get(grupo.id, 0)
        avg_pct = 0.0
        if total_sesiones and total_participantes > 0:
            pcts = [
                present / total_participantes * 100
                for present in present_by_group_session.get(grupo.id, {}).values()
            ]
            avg_pct = round(sum(pcts) / len(pcts), 1)

        grupos_resumen.append({
            "grupo_id": grupo.id,
            "nombre": grupo.nombre,
            "codigo": grupo.codigo,
            "lider": leader_name,
            "participantes": total_participantes,
            "total_sesiones": total_sesiones,
            "sesiones_realizadas": sesiones_realizadas,
            "promedio_asistencia_pct": avg_pct,
        })

    return {
        "estrategia_id": estrategia.id,
        "nombre": estrategia.nombre,
        "categoria_id": estrategia.categoria_id,
        "frecuencia": estrategia.frecuencia,
        "activa": estrategia.activa,
        "total_grupos": len(grupos_resumen),
        "grupos": grupos_resumen,
    }
